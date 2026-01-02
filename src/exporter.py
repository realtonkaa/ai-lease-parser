"""Export extracted lease data to CSV or Excel."""

import csv
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from .fields import get_all_field_names


def _flatten_value(value: Any) -> str:
    """Convert a field value to a CSV-friendly string."""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return str(value)


def _rows_to_flat(records: List[Dict[str, Any]]) -> Tuple:
    """Convert a list of record dicts to (headers, rows) for export."""
    headers = get_all_field_names()
    rows = []
    for record in records:
        row = [_flatten_value(record.get(field)) for field in headers]
        rows.append(row)
    return headers, rows


def export_csv(
    records: List[Dict[str, Any]],
    output_path: str,
    source_column: bool = True,
) -> str:
    """
    Export lease records to a CSV file.

    Args:
        records: List of dicts mapping field_name -> value.
                 May optionally include a "_source" key for the filename.
        output_path: Path to write the CSV.
        source_column: If True, include a "source_file" column.

    Returns:
        Absolute path to written CSV.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = get_all_field_names()
    if source_column:
        headers = ["source_file"] + headers

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for record in records:
            if source_column:
                row = [_flatten_value(record.get("_source", ""))]
            else:
                row = []
            row += [_flatten_value(record.get(field)) for field in get_all_field_names()]
            writer.writerow(row)

    return str(output_path.resolve())


def export_excel(
    records: List[Dict[str, Any]],
    output_path: str,
    sheet_name: str = "Leases",
    source_column: bool = True,
) -> str:
    """
    Export lease records to an Excel (.xlsx) file.

    Args:
        records: List of dicts mapping field_name -> value.
        output_path: Path to write the Excel file.
        sheet_name: Name of the worksheet.
        source_column: If True, include a "source_file" column.

    Returns:
        Absolute path to written Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = get_all_field_names()
    if source_column:
        headers = ["source_file"] + headers

    # Write header row with bold formatting
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        if source_column:
            row = [_flatten_value(record.get("_source", ""))]
        else:
            row = []
        row += [_flatten_value(record.get(field)) for field in get_all_field_names()]

        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(str(output_path))
    return str(output_path.resolve())


def export(
    records: List[Dict[str, Any]],
    output_path: str,
    source_column: bool = True,
) -> str:
    """
    Export lease records to CSV or Excel based on file extension.

    Args:
        records: List of extracted lease data dicts.
        output_path: Output file path (.csv or .xlsx).
        source_column: Include source filename column.

    Returns:
        Absolute path to written file.
    """
    ext = Path(output_path).suffix.lower()
    if ext == ".csv":
        return export_csv(records, output_path, source_column=source_column)
    elif ext in (".xlsx", ".xls"):
        return export_excel(records, output_path, source_column=source_column)
    else:
        raise ValueError(f"Unsupported output format: {ext}. Use .csv or .xlsx")


