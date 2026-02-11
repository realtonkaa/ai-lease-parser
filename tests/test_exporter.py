"""Tests for src/exporter.py — CSV and Excel export."""

import csv
import os
from pathlib import Path
import pytest

from src.exporter import export_csv, export_excel, export, _flatten_value


class TestFlattenValue:
    """Tests for the _flatten_value helper."""

    def test_none_becomes_empty_string(self):
        assert _flatten_value(None) == ""

    def test_float_becomes_string(self):
        assert _flatten_value(1450.0) == "1450.0"

    def test_list_becomes_semicolon_separated(self):
        assert _flatten_value(["water", "gas"]) == "water; gas"

    def test_string_passthrough(self):
        assert _flatten_value("hello") == "hello"

    def test_empty_list_becomes_empty_string(self):
        assert _flatten_value([]) == ""

    def test_integer_becomes_string(self):
        assert _flatten_value(5) == "5"


class TestExportCSV:
    """Tests for export_csv()."""

    def test_creates_file(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "test.csv")
        result = export_csv([full_values], out)
        assert Path(result).exists()

    def test_returns_absolute_path(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "test.csv")
        result = export_csv([full_values], out)
        assert os.path.isabs(result)

    def test_csv_has_header_row(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "headers.csv")
        export_csv([full_values], out)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert "monthly_rent" in header
        assert "tenant_name" in header

    def test_csv_has_data_row(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "data.csv")
        export_csv([full_values], out)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 2  # header + 1 data row

    def test_csv_multiple_records(self, tmp_output_dir, full_values, minimal_valid_values):
        out = str(tmp_output_dir / "multi.csv")
        export_csv([full_values, minimal_valid_values], out)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 3  # header + 2 data rows

    def test_csv_source_column_included(self, tmp_output_dir, full_values):
        full_values["_source"] = "lease1.pdf"
        out = str(tmp_output_dir / "source.csv")
        export_csv([full_values], out, source_column=True)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert "source_file" in header

    def test_csv_source_column_excluded(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "nosource.csv")
        export_csv([full_values], out, source_column=False)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert "source_file" not in header

    def test_csv_values_correct(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "values.csv")
        export_csv([full_values], out, source_column=False)
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            row = next(reader)
        assert row["monthly_rent"] == "1450.0"

    def test_csv_empty_records(self, tmp_output_dir):
        out = str(tmp_output_dir / "empty.csv")
        result = export_csv([], out)
        assert Path(result).exists()
        with open(out, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 1  # Just the header

    def test_csv_creates_parent_dirs(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "subdir" / "nested" / "test.csv")
        result = export_csv([full_values], out)
        assert Path(result).exists()


class TestExportExcel:
    """Tests for export_excel()."""

    def test_creates_xlsx_file(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "test.xlsx")
        result = export_excel([full_values], out)
        assert Path(result).exists()

    def test_xlsx_has_correct_extension(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "test.xlsx")
        result = export_excel([full_values], out)
        assert result.endswith(".xlsx")

    def test_xlsx_readable(self, tmp_output_dir, full_values):
        import openpyxl
        out = str(tmp_output_dir / "readable.xlsx")
        export_excel([full_values], out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row >= 2

    def test_xlsx_header_row(self, tmp_output_dir, full_values):
        import openpyxl
        out = str(tmp_output_dir / "header.xlsx")
        export_excel([full_values], out, source_column=False)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "monthly_rent" in headers

    def test_xlsx_sheet_name(self, tmp_output_dir, full_values):
        import openpyxl
        out = str(tmp_output_dir / "sheet.xlsx")
        export_excel([full_values], out, sheet_name="TestSheet")
        wb = openpyxl.load_workbook(out)
        assert "TestSheet" in wb.sheetnames

    def test_xlsx_multiple_records(self, tmp_output_dir, full_values, minimal_valid_values):
        import openpyxl
        out = str(tmp_output_dir / "multi.xlsx")
        export_excel([full_values, minimal_valid_values], out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows


class TestExportDispatcher:
    """Tests for the export() dispatcher function."""

    def test_csv_extension_routes_to_csv(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "out.csv")
        result = export([full_values], out)
        assert result.endswith(".csv")
        assert Path(result).exists()

    def test_xlsx_extension_routes_to_excel(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "out.xlsx")
        result = export([full_values], out)
        assert result.endswith(".xlsx")
        assert Path(result).exists()

    def test_unsupported_extension_raises(self, tmp_output_dir, full_values):
        out = str(tmp_output_dir / "out.json")
        with pytest.raises(ValueError, match="Unsupported output format"):
            export([full_values], out)
